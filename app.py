"""
GHF - Sistema de Controle de Contratos de Experiencia
Backend Principal - Flask Application
"""

import os
import sys
import hashlib
import secrets
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, send_file, flash, abort)
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None

# ==================== CONFIGURACAO ====================
app = Flask(__name__, static_folder='static', static_url_path='/static')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Detectar se esta no Render.com (online) ou local
DATABASE_URL = os.environ.get('DATABASE_URL')
USANDO_POSTGRES = DATABASE_URL is not None and DATABASE_URL.startswith('postgres')

if USANDO_POSTGRES:
    # Producao online - usa PostgreSQL
    import psycopg2
    import psycopg2.extras
    DATABASE = DATABASE_URL
    print("✅ Conectando ao PostgreSQL (Render.com)")
else:
    # Desenvolvimento local - usa SQLite
    DATABASE = os.path.join(BASE_DIR, 'dados', 'contratos.db')
    print("✅ Conectando ao SQLite (local)")

EXCEL_FILE = os.path.join(BASE_DIR, 'dados', 'CONTRATO DE EXPERIENCIA.xlsx')

# Chave secreta fixa para manter sessoes entre reinicializacoes
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    SECRET_FILE = os.path.join(BASE_DIR, 'dados', '.secret_key')
    if os.path.exists(SECRET_FILE):
        with open(SECRET_FILE, 'r') as f:
            SECRET_KEY = f.read().strip()
    else:
        SECRET_KEY = secrets.token_hex(32)
        os.makedirs(os.path.dirname(SECRET_FILE), exist_ok=True)
        with open(SECRET_FILE, 'w') as f:
            f.write(SECRET_KEY)

app.secret_key = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

# MIME types para PWA
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.after_request
def add_header(response):
    # Service Worker precisa de headers especificos
    if request.path.endswith('service-worker.js'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Service-Worker-Allowed'] = '/'
    # Manifest JSON
    if request.path.endswith('manifest.json'):
        response.headers['Content-Type'] = 'application/manifest+json'
    return response

# ==================== CONEXAO BANCO ====================
def get_db():
    if USANDO_POSTGRES:
        conn = psycopg2.connect(DATABASE)
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        return conn
    else:
        import sqlite3
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    
    if USANDO_POSTGRES:
        # PostgreSQL
        cur.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                senha_hash TEXT NOT NULL,
                perfil TEXT NOT NULL CHECK(perfil IN ('dp', 'gestor', 'regional')),
                loja TEXT,
                regional TEXT,
                ativo INTEGER DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS colaboradores (
                id SERIAL PRIMARY KEY,
                empresa TEXT,
                nome_empresa TEXT,
                matricula TEXT,
                nome TEXT NOT NULL,
                ddd TEXT,
                celular TEXT,
                funcao TEXT,
                email TEXT,
                data_admissao TEXT,
                vencimento_1 TEXT,
                vencimento_2 TEXT,
                departamento TEXT,
                setor TEXT,
                status_1 TEXT DEFAULT 'pendente',
                status_2 TEXT DEFAULT 'pendente',
                validacao_gestor_1 TEXT,
                validacao_data_1 TEXT,
                validacao_gestor_2 TEXT,
                validacao_data_2 TEXT,
                alertas_enviados INTEGER DEFAULT 0,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS alertas_log (
                id SERIAL PRIMARY KEY,
                colaborador_id INTEGER,
                tipo TEXT,
                periodo INTEGER,
                data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (colaborador_id) REFERENCES colaboradores(id)
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS gestores (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                cargo TEXT NOT NULL,
                loja TEXT,
                lojas TEXT,
                regional TEXT,
                ddd TEXT,
                celular TEXT,
                email TEXT,
                ativo INTEGER DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS config_sistema (
                chave TEXT PRIMARY KEY,
                valor TEXT
            )
        ''')
    else:
        # SQLite
        os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                senha_hash TEXT NOT NULL,
                perfil TEXT NOT NULL CHECK(perfil IN ('dp', 'gestor', 'regional')),
                loja TEXT,
                regional TEXT,
                ativo INTEGER DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS colaboradores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                empresa TEXT,
                nome_empresa TEXT,
                matricula TEXT,
                nome TEXT NOT NULL,
                ddd TEXT,
                celular TEXT,
                funcao TEXT,
                email TEXT,
                data_admissao TEXT,
                vencimento_1 TEXT,
                vencimento_2 TEXT,
                departamento TEXT,
                setor TEXT,
                status_1 TEXT DEFAULT 'pendente' CHECK(status_1 IN ('pendente', 'renovado', 'desligado')),
                status_2 TEXT DEFAULT 'pendente' CHECK(status_2 IN ('pendente', 'efetivado', 'desligado')),
                validacao_gestor_1 TEXT,
                validacao_data_1 TEXT,
                validacao_gestor_2 TEXT,
                validacao_data_2 TEXT,
                alertas_enviados INTEGER DEFAULT 0,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS alertas_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                colaborador_id INTEGER,
                tipo TEXT CHECK(tipo IN ('email', 'whatsapp')),
                periodo INTEGER CHECK(periodo IN (1, 2)),
                data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (colaborador_id) REFERENCES colaboradores(id)
            );

            CREATE TABLE IF NOT EXISTS config_sistema (
                chave TEXT PRIMARY KEY,
                valor TEXT
            );
        ''')
    
    conn.commit()
    cur.close()
    conn.close()

# ==================== UTILIDADES ====================
def hash_senha(senha):
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()

def converter_data_para_db(data_str):
    if not data_str or data_str.strip() == '':
        return None
    try:
        return datetime.strptime(data_str.strip(), '%d/%m/%Y').strftime('%Y-%m-%d')
    except:
        try:
            return datetime.strptime(data_str.strip(), '%Y-%m-%d').strftime('%Y-%m-%d')
        except:
            return None

def formatar_data(data_str):
    if not data_str:
        return '-'
    try:
        return datetime.strptime(data_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        return data_str

def calcular_dias_restantes(data_str):
    if not data_str:
        return None
    try:
        vencimento = datetime.strptime(data_str, '%Y-%m-%d').date()
        hoje = datetime.now().date()
        return (vencimento - hoje).days
    except:
        return None

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Acesso negado. Faca login.', 'erro')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def dp_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('perfil') != 'dp':
            flash('Acesso restrito ao DP.', 'erro')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== ROTAS AUTENTICACAO ====================
@app.route('/')
def index():
    if 'usuario_id' in session:
        if session.get('perfil') == 'dp':
            return redirect(url_for('dp_dashboard'))
        else:
            return redirect(url_for('gestor_dashboard'))
    return redirect(url_for('login'))

@app.route('/acesso')
def acesso():
    """Pagina publica com QR Code e instrucoes de instalacao"""
    return render_template('acesso.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '').strip()

        if not email or not senha:
            flash('Preencha todos os campos.', 'erro')
            return render_template('login.html')

        db = get_db()
        usuario = db.execute(
            'SELECT * FROM usuarios WHERE email = ? AND senha_hash = ? AND ativo = 1',
            (email, hash_senha(senha))
        ).fetchone()
        db.close()

        if usuario:
            session.permanent = True
            session['usuario_id'] = usuario['id']
            session['usuario_nome'] = usuario['nome']
            session['perfil'] = usuario['perfil']
            session['loja'] = usuario['loja']
            session['lojas'] = usuario['lojas'] if usuario['lojas'] else ''
            session['regional'] = usuario['regional'] if usuario['regional'] else ''

            if usuario['perfil'] == 'dp':
                return redirect(url_for('dp_dashboard'))
            else:
                return redirect(url_for('gestor_dashboard'))

        flash('E-mail ou senha invalidos.', 'erro')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso.', 'sucesso')
    return redirect(url_for('login'))

# ==================== PAINEL DP ====================
@app.route('/dp')
@login_required
@dp_required
def dp_dashboard():
    db = get_db()
    hoje = datetime.now().date().isoformat()

    # Filtros
    filtro_regional = request.args.get('regional', '').strip()
    filtro_loja = request.args.get('loja', '').strip()
    filtro_status = request.args.get('status', '').strip()
    filtro_busca = request.args.get('busca', '').strip()

    query = 'SELECT * FROM colaboradores WHERE 1=1'
    params = []

    if filtro_regional:
        query += ' AND setor = ?'
        params.append(filtro_regional)
    if filtro_loja:
        query += ' AND departamento = ?'
        params.append(filtro_loja)
    if filtro_status == 'pendente':
        query += ' AND (status_1 = "pendente" OR status_2 = "pendente")'
    elif filtro_status == 'atrasado':
        query += ' AND ((status_1 = "pendente" AND vencimento_1 <= ?) OR (status_2 = "pendente" AND vencimento_2 <= ?))'
        params.extend([hoje, hoje])
    elif filtro_status == 'concluido':
        query += ' AND status_1 != "pendente" AND status_2 != "pendente"'
    elif filtro_status == 'periodo1':
        query += ' AND status_1 = "pendente"'
    elif filtro_status == 'periodo2':
        query += ' AND status_2 = "pendente" AND status_1 != "pendente"'

    if filtro_busca:
        query += ' AND (nome LIKE ? OR matricula LIKE ? OR funcao LIKE ?)'
        busca = f'%{filtro_busca}%'
        params.extend([busca, busca, busca])

    query += ' ORDER BY setor, departamento, nome'
    colaboradores = db.execute(query, params).fetchall()

    # Estatisticas
    stats = db.execute('''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status_1 = 'pendente' OR status_2 = 'pendente' THEN 1 ELSE 0 END) as pendentes,
            SUM(CASE WHEN (status_1 = 'pendente' AND vencimento_1 <= ?) OR (status_2 = 'pendente' AND vencimento_2 <= ?) THEN 1 ELSE 0 END) as atrasados,
            SUM(CASE WHEN status_1 = 'renovado' THEN 1 ELSE 0 END) as renovados,
            SUM(CASE WHEN status_2 = 'efetivado' THEN 1 ELSE 0 END) as efetivados,
            SUM(CASE WHEN status_1 = 'desligado' OR status_2 = 'desligado' THEN 1 ELSE 0 END) as desligados
        FROM colaboradores
    ''', (hoje, hoje)).fetchone()

    # Listas para filtros
    regionais = db.execute(
        'SELECT DISTINCT setor FROM colaboradores WHERE setor IS NOT NULL AND setor != "" ORDER BY setor'
    ).fetchall()
    lojas = db.execute(
        'SELECT DISTINCT departamento FROM colaboradores WHERE departamento IS NOT NULL AND departamento != "" ORDER BY departamento'
    ).fetchall()

    # Gestores para contatos WhatsApp
    gestores = db.execute('SELECT * FROM gestores WHERE ativo = 1').fetchall()
    gestores_map = {}
    for g in gestores:
        if g['regional']:
            gestores_map[g['regional'].upper()] = dict(g)
        if g['lojas']:
            for l in g['lojas'].split(','):
                gestores_map[l.strip().upper()] = dict(g)
        if g['loja']:
            gestores_map[g['loja'].upper()] = dict(g)

    # Enriquecer colaboradores com dados do gestor e dias restantes
    colaboradores_list = []
    for c in colaboradores:
        item = dict(c)
        gestor = gestores_map.get(c['departamento'].upper()) or gestores_map.get(c['setor'].upper(), None)
        item['gestor'] = gestor
        item['dias_restantes_1'] = calcular_dias_restantes(c['vencimento_1'])
        item['dias_restantes_2'] = calcular_dias_restantes(c['vencimento_2'])
        colaboradores_list.append(item)

    db.close()

    # URL do sistema para links
    config_url = request.host_url.rstrip('/')

    return render_template('dp_dashboard.html',
        colaboradores=colaboradores_list,
        stats=stats,
        regionais=regionais,
        lojas=lojas,
        filtros={
            'regional': filtro_regional,
            'loja': filtro_loja,
            'status': filtro_status,
            'busca': filtro_busca
        },
        config_url=config_url)

# ==================== PAINEL GESTOR ====================
@app.route('/gestor')
@login_required
def gestor_dashboard():
    db = get_db()
    hoje = datetime.now().date().isoformat()

    perfil = session.get('perfil')
    gestor_loja = session.get('loja')
    gestor_lojas = session.get('lojas')
    gestor_regional = session.get('regional')

    if perfil == 'dp':
        # DP pode ver tudo
        query = 'SELECT * FROM colaboradores WHERE 1=1'
        params = []
    elif perfil == 'supervisor' and gestor_regional:
        # Supervisor ve toda a regional
        query = 'SELECT * FROM colaboradores WHERE setor = ?'
        params = [gestor_regional]
    elif perfil == 'gerente_area' and gestor_lojas:
        # Gerente de area ve varias lojas (separadas por virgula)
        lojas_lista = [l.strip() for l in gestor_lojas.split(',')]
        placeholders = ','.join(['?' for _ in lojas_lista])
        query = f'SELECT * FROM colaboradores WHERE departamento IN ({placeholders})'
        params = lojas_lista
    elif perfil == 'subgerente' and gestor_loja:
        # Subgerente ve apenas sua loja
        query = 'SELECT * FROM colaboradores WHERE departamento = ?'
        params = [gestor_loja]
    else:
        query = 'SELECT * WHERE 1=0'
        params = []

    colaboradores = db.execute(query, params).fetchall()

    # Separar por periodo
    periodo_1_pendente = []
    periodo_2_pendente = []
    historico = []

    for c in colaboradores:
        dias1 = calcular_dias_restantes(c['vencimento_1'])
        dias2 = calcular_dias_restantes(c['vencimento_2'])

        item = dict(c)
        item['dias_restantes_1'] = dias1
        item['dias_restantes_2'] = dias2
        item['vencimento_formatado_1'] = formatar_data(c['vencimento_1']) if c['vencimento_1'] else '-'
        item['vencimento_formatado_2'] = formatar_data(c['vencimento_2']) if c['vencimento_2'] else '-'

        if c['status_1'] == 'pendente' and c['vencimento_1']:
            periodo_1_pendente.append(item)
        elif c['status_2'] == 'pendente' and c['vencimento_2'] and c['status_1'] != 'pendente':
            periodo_2_pendente.append(item)
        elif c['status_1'] != 'pendente' or c['status_2'] != 'pendente':
            historico.append(item)

    # Ordenar por urgencia
    periodo_1_pendente.sort(key=lambda x: x['dias_restantes_1'] if x['dias_restantes_1'] is not None else 999)
    periodo_2_pendente.sort(key=lambda x: x['dias_restantes_2'] if x['dias_restantes_2'] is not None else 999)

    db.close()

    # URL do sistema para links
    config_url = request.host_url.rstrip('/')

    return render_template('gestor_dashboard.html',
        periodo_1=periodo_1_pendente,
        periodo_2=periodo_2_pendente,
        historico=historico,
        hoje=datetime.now().date(),
        total_p1=len(periodo_1_pendente),
        total_p2=len(periodo_2_pendente),
        total_hist=len(historico),
        config_url=config_url)

# ==================== DASHBOARD SUPERVISOR ====================
@app.route('/supervisor')
@login_required
def supervisor_dashboard():
    db = get_db()
    hoje = datetime.now().date().isoformat()

    # Regional do gestor logado
    regional_gestor = session.get('regional', '')
    perfil = session.get('perfil', '')

    # Filtro de loja (opcional)
    filtro_loja = request.args.get('loja', '').strip()

    # Condicao WHERE - sempre filtra pela regional do gestor
    where_extra = ''
    params_extra = []
    if regional_gestor:
        where_extra += ' AND setor = ?'
        params_extra.append(regional_gestor)
    if filtro_loja:
        where_extra += ' AND departamento = ?'
        params_extra.append(filtro_loja)

    # Estatisticas gerais da regional
    stats = db.execute(f'''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status_1 = 'pendente' OR status_2 = 'pendente' THEN 1 ELSE 0 END) as pendentes,
            SUM(CASE WHEN (status_1 = 'pendente' AND vencimento_1 <= ?) OR (status_2 = 'pendente' AND vencimento_2 <= ?) THEN 1 ELSE 0 END) as atrasados,
            SUM(CASE WHEN status_1 = 'renovado' THEN 1 ELSE 0 END) as renovados,
            SUM(CASE WHEN status_2 = 'efetivado' THEN 1 ELSE 0 END) as efetivados,
            SUM(CASE WHEN status_1 = 'desligado' OR status_2 = 'desligado' THEN 1 ELSE 0 END) as desligados
        FROM colaboradores WHERE 1=1 {where_extra}
    ''', (hoje, hoje) + tuple(params_extra)).fetchone()

    # Estatisticas por loja da regional (visao de gestao)
    lojas_stats = db.execute(f'''
        SELECT
            departamento as loja,
            COUNT(*) as total,
            SUM(CASE WHEN status_1 = 'pendente' OR status_2 = 'pendente' THEN 1 ELSE 0 END) as em_experiencia,
            SUM(CASE WHEN (status_1 = 'pendente' AND vencimento_1 <= ?) OR (status_2 = 'pendente' AND vencimento_2 <= ?) THEN 1 ELSE 0 END) as atrasados,
            SUM(CASE WHEN status_1 = 'renovado' THEN 1 ELSE 0 END) as renovados,
            SUM(CASE WHEN status_2 = 'efetivado' THEN 1 ELSE 0 END) as efetivados,
            SUM(CASE WHEN status_1 = 'desligado' OR status_2 = 'desligado' THEN 1 ELSE 0 END) as desligados
        FROM colaboradores
        WHERE departamento IS NOT NULL AND departamento != '' {where_extra}
        GROUP BY departamento
        ORDER BY departamento
    ''', (hoje, hoje) + tuple(params_extra)).fetchall()

    # Calcular efetivos (fora do periodo de experiencia) por loja
    lojas_stats_list = []
    for loja in lojas_stats:
        item = dict(loja)
        total = item['total']
        em_exp = item['em_experiencia'] if item['em_experiencia'] else 0
        efet = item['efetivados'] if item['efetivados'] else 0
        ren = item['renovados'] if item['renovados'] else 0
        desl = item['desligados'] if item['desligados'] else 0
        item['efetivos'] = total - em_exp - desl
        lojas_stats_list.append(item)
    lojas_stats = lojas_stats_list

    # Colaboradores atrasados (urgente)
    atrasados_raw = db.execute(f'''
        SELECT c.*, 
            CASE 
                WHEN status_1 = 'pendente' AND vencimento_1 <= ? THEN 1
                WHEN status_2 = 'pendente' AND vencimento_2 <= ? THEN 2
            END as periodo_atrasado,
            CASE 
                WHEN status_1 = 'pendente' AND vencimento_1 <= ? THEN vencimento_1
                WHEN status_2 = 'pendente' AND vencimento_2 <= ? THEN vencimento_2
            END as vencimento_atrasado
        FROM colaboradores c
        WHERE (status_1 = 'pendente' AND vencimento_1 <= ?) 
           OR (status_2 = 'pendente' AND vencimento_2 <= ?)
        {where_extra}
        ORDER BY vencimento_atrasado ASC
    ''', (hoje, hoje, hoje, hoje, hoje, hoje) + tuple(params_extra)).fetchall()

    atrasados = []
    hoje_date = datetime.now().date()
    for a in atrasados_raw:
        item = dict(a)
        if a['vencimento_atrasado']:
            venc_date = datetime.strptime(a['vencimento_atrasado'][:10], '%Y-%m-%d').date()
            item['dias_atrasado'] = (hoje_date - venc_date).days
        else:
            item['dias_atrasado'] = 0
        atrasados.append(item)

    # Colaboradores pendentes (precisam de validacao)
    pendentes_raw = db.execute(f'''
        SELECT c.*,
            CASE 
                WHEN status_1 = 'pendente' THEN vencimento_1
                ELSE vencimento_2
            END as vencimento_pendente,
            CASE 
                WHEN status_1 = 'pendente' THEN 1
                ELSE 2
            END as periodo_pendente
        FROM colaboradores c
        WHERE (status_1 = 'pendente' OR (status_2 = 'pendente' AND status_1 != 'pendente'))
        {where_extra}
        ORDER BY vencimento_pendente ASC
    ''', tuple(params_extra)).fetchall()

    pendentes = []
    for p in pendentes_raw:
        item = dict(p)
        if p['vencimento_pendente']:
            venc_date = datetime.strptime(p['vencimento_pendente'][:10], '%Y-%m-%d').date()
            dias = (venc_date - hoje_date).days
            item['dias_restantes'] = dias
            if dias <= 0:
                item['urgencia'] = 'vencido'
            elif dias <= 3:
                item['urgencia'] = 'urgente'
            elif dias <= 7:
                item['urgencia'] = 'atencao'
            else:
                item['urgencia'] = 'ok'
        else:
            item['dias_restantes'] = 0
            item['urgencia'] = 'ok'
        pendentes.append(item)

    # Lista de lojas para filtro
    lojas = db.execute(
        f'SELECT DISTINCT departamento FROM colaboradores WHERE departamento IS NOT NULL AND departamento != "" {where_extra} ORDER BY departamento',
        tuple(params_extra)
    ).fetchall()

    db.close()

    return render_template('supervisor_dashboard.html',
        stats=stats,
        lojas_stats=lojas_stats,
        atrasados=atrasados,
        pendentes=pendentes,
        lojas=lojas,
        filtros={'loja': filtro_loja},
        regional=regional_gestor,
        hoje=datetime.now().date())

# ==================== VALIDACAO GESTOR ====================
@app.route('/validar', methods=['POST'])
@login_required
def validar():
    db = get_db()
    try:
        dados = request.get_json()
        colaborador_id = dados.get('colaborador_id')
        periodo = dados.get('periodo')
        acao = dados.get('acao')

        if periodo not in [1, 2] or acao not in ['renovado', 'efetivado', 'desligado']:
            return jsonify({'sucesso': False, 'mensagem': 'Parametros invalidos'}), 400

        if periodo == 1 and acao not in ['renovado', 'desligado']:
            return jsonify({'sucesso': False, 'mensagem': 'Acao invalida para periodo 1'}), 400

        if periodo == 2 and acao not in ['efetivado', 'desligado']:
            return jsonify({'sucesso': False, 'mensagem': 'Acao invalida para periodo 2'}), 400

        col = db.execute('SELECT * FROM colaboradores WHERE id = ?', (colaborador_id,)).fetchone()
        if not col:
            return jsonify({'sucesso': False, 'mensagem': 'Colaborador nao encontrado'}), 404

        # Verificar se gestor tem permissao
        if session['perfil'] != 'dp':
            if session.get('loja') and col['departamento'] != session['loja']:
                return jsonify({'sucesso': False, 'mensagem': 'Sem permissao'}), 403

        now = datetime.now().isoformat()

        if periodo == 1:
            db.execute('''
                UPDATE colaboradores
                SET status_1 = ?, validacao_gestor_1 = ?, validacao_data_1 = ?, atualizado_em = ?
                WHERE id = ?
            ''', (acao, session['usuario_nome'], now, now, colaborador_id))
        else:
            db.execute('''
                UPDATE colaboradores
                SET status_2 = ?, validacao_gestor_2 = ?, validacao_data_2 = ?, atualizado_em = ?
                WHERE id = ?
            ''', (acao, session['usuario_nome'], now, now, colaborador_id))

        db.commit()
        return jsonify({'sucesso': True, 'mensagem': f'Registrado: {acao}'})

    except Exception as e:
        db.rollback()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500
    finally:
        db.close()

# ==================== EXPORTAR EXCEL ====================
@app.route('/exportar')
@login_required
def exportar():
    if not openpyxl:
        flash('Biblioteca openpyxl nao instalada.', 'erro')
        return redirect(url_for('dp_dashboard'))

    db = get_db()

    if session['perfil'] == 'dp':
        colaboradores = db.execute('SELECT * FROM colaboradores ORDER BY setor, departamento, nome').fetchall()
    elif session.get('loja'):
        colaboradores = db.execute('SELECT * FROM colaboradores WHERE departamento = ? ORDER BY nome', (session['loja'],)).fetchall()
    else:
        colaboradores = []

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos Experiencia"

    headers = ['Empresa', 'Nome Empresa', 'Matricula', 'Nome', 'DDD', 'Celular',
               'Funcao', 'Email', 'Admissao', 'Vencimento 1', 'Vencimento 2',
               'Departamento', 'Setor', 'Status 1', 'Status 2',
               'Validado por (P1)', 'Data Validacao P1',
               'Validado por (P2)', 'Data Validacao P2']

    # Estilizar cabecalhos
    from openpyxl.styles import Font, PatternFill, Alignment
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    for row_idx, c in enumerate(colaboradores, 2):
        ws.cell(row=row_idx, column=1, value=c['empresa'])
        ws.cell(row=row_idx, column=2, value=c['nome_empresa'])
        ws.cell(row=row_idx, column=3, value=c['matricula'])
        ws.cell(row=row_idx, column=4, value=c['nome'])
        ws.cell(row=row_idx, column=5, value=c['ddd'])
        ws.cell(row=row_idx, column=6, value=c['celular'])
        ws.cell(row=row_idx, column=7, value=c['funcao'])
        ws.cell(row=row_idx, column=8, value=c['email'])
        ws.cell(row=row_idx, column=9, value=formatar_data(c['data_admissao']))
        ws.cell(row=row_idx, column=10, value=formatar_data(c['vencimento_1']))
        ws.cell(row=row_idx, column=11, value=formatar_data(c['vencimento_2']))
        ws.cell(row=row_idx, column=12, value=c['departamento'])
        ws.cell(row=row_idx, column=13, value=c['setor'])
        ws.cell(row=row_idx, column=14, value=c['status_1'])
        ws.cell(row=row_idx, column=15, value=c['status_2'])
        ws.cell(row=row_idx, column=16, value=c['validacao_gestor_1'])
        ws.cell(row=row_idx, column=17, value=c['validacao_data_1'])
        ws.cell(row=row_idx, column=18, value=c['validacao_gestor_2'])
        ws.cell(row=row_idx, column=19, value=c['validacao_data_2'])

        # Colorir status
        fill_verde = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        fill_vermelho = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        if c['status_1'] == 'renovado' or c['status_2'] == 'efetivado':
            ws.cell(row=row_idx, column=14).fill = fill_verde
            ws.cell(row=row_idx, column=15).fill = fill_verde
        elif c['status_1'] == 'desligado' or c['status_2'] == 'desligado':
            ws.cell(row=row_idx, column=14).fill = fill_vermelho
            ws.cell(row=row_idx, column=15).fill = fill_vermelho

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'contratos_experiencia_{timestamp}.xlsx')

# ==================== GERENCIAMENTO DE USUARIOS ====================
@app.route('/usuarios')
@login_required
@dp_required
def gerenciar_usuarios():
    db = get_db()
    usuarios = db.execute('SELECT * FROM usuarios ORDER BY perfil, nome').fetchall()
    db.close()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuarios/novo', methods=['POST'])
@login_required
@dp_required
def novo_usuario():
    db = get_db()
    try:
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '').strip()
        perfil = request.form.get('perfil', 'gestor')
        loja = request.form.get('loja', '').strip()
        regional = request.form.get('regional', '').strip()

        if not nome or not email or not senha:
            flash('Preencha nome, e-mail e senha.', 'erro')
            return redirect(url_for('gerenciar_usuarios'))

        db.execute('''
            INSERT INTO usuarios (nome, email, senha_hash, perfil, loja, regional)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (nome, email, hash_senha(senha), perfil, loja or None, regional or None))
        db.commit()
        flash('Usuario criado com sucesso!', 'sucesso')
    except sqlite3.IntegrityError:
        flash('E-mail ja cadastrado.', 'erro')
    except Exception as e:
        flash(f'Erro: {e}', 'erro')
    finally:
        db.close()
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/usuarios/toggle/<int:usuario_id>')
@login_required
@dp_required
def toggle_usuario(usuario_id):
    db = get_db()
    db.execute('UPDATE usuarios SET ativo = CASE WHEN ativo = 1 THEN 0 ELSE 1 END WHERE id = ?', (usuario_id,))
    db.commit()
    db.close()
    flash('Status do usuario alterado.', 'sucesso')
    return redirect(url_for('gerenciar_usuarios'))

# ==================== ATUALIZAR DADOS ====================
@app.route('/atualizar')
@login_required
@dp_required
def atualizar_pagina():
    db = get_db()
    hoje = datetime.now().date().isoformat()

    stats = db.execute('''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status_1 != 'pendente' OR status_2 != 'pendente' THEN 1 ELSE 0 END) as validados
        FROM colaboradores
    ''').fetchone()

    ultima = db.execute('SELECT MAX(atualizado_em) as data FROM colaboradores').fetchone()

    db.close()

    return render_template('atualizar.html',
        stats={
            'total': stats['total'] or 0,
            'pendentes': (stats['total'] or 0) - (stats['validados'] or 0),
            'validados': stats['validados'] or 0,
            'ultima_atualizacao': ultima['data'][:10] if ultima and ultima['data'] else 'Nunca'
        })

@app.route('/atualizar/upload', methods=['POST'])
@login_required
@dp_required
def upload_planilha():
    if 'arquivo' not in request.files:
        flash('Nenhum arquivo selecionado.', 'erro')
        return redirect(url_for('atualizar_pagina'))

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        flash('Nenhum arquivo selecionado.', 'erro')
        return redirect(url_for('atualizar_pagina'))

    if not arquivo.filename.lower().endswith(('.xls', '.xlsx')):
        flash('Formato invalido. Use .xls ou .xlsx', 'erro')
        return redirect(url_for('atualizar_pagina'))

    # Salvar arquivo
    caminho_dados = os.path.join(BASE_DIR, 'dados')
    os.makedirs(caminho_dados, exist_ok=True)

    # Manter backup do arquivo antigo
    arquivo_antigo = os.path.join(caminho_dados, 'CONTRATO DE EXPERIENCIA.xlsx')
    if os.path.exists(arquivo_antigo):
        backup = os.path.join(caminho_dados, f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        os.rename(arquivo_antigo, backup)

    # Salvar novo arquivo
    caminho_novo = os.path.join(caminho_dados, arquivo.filename)
    arquivo.save(caminho_novo)

    # Importar dados
    try:
        sys.path.insert(0, BASE_DIR)
        from database import importar_excel, criar_usuario_dp
        importar_excel(caminho_novo)
        criar_usuario_dp()
        flash(f'Planilha "{arquivo.filename}" importada com sucesso!', 'sucesso')
    except Exception as e:
        flash(f'Erro ao importar: {e}', 'erro')

    return redirect(url_for('atualizar_pagina'))

# ==================== API ====================
@app.route('/api/stats')
@login_required
def api_stats():
    db = get_db()
    hoje = datetime.now().date().isoformat()

    stats = db.execute('''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status_1 = 'pendente' THEN 1 ELSE 0 END) as p1_pendente,
            SUM(CASE WHEN status_2 = 'pendente' AND status_1 != 'pendente' THEN 1 ELSE 0 END) as p2_pendente,
            SUM(CASE WHEN (status_1 = 'pendente' AND vencimento_1 <= ?) THEN 1 ELSE 0 END) as p1_atrasado,
            SUM(CASE WHEN (status_2 = 'pendente' AND vencimento_2 <= ?) THEN 1 ELSE 0 END) as p2_atrasado,
            SUM(CASE WHEN status_1 = 'renovado' THEN 1 ELSE 0 END) as renovados,
            SUM(CASE WHEN status_2 = 'efetivado' THEN 1 ELSE 0 END) as efetivados,
            SUM(CASE WHEN status_1 = 'desligado' OR status_2 = 'desligado' THEN 1 ELSE 0 END) as desligados
        FROM colaboradores
    ''', (hoje, hoje)).fetchone()

    db.close()
    return jsonify(dict(stats))

# ==================== INICIAR ====================
if __name__ == '__main__':
    init_db()

    print("=" * 50)
    print("  GHF - Sistema de Controle de Experiencia")
    print("=" * 50)
    print()

    if USANDO_POSTGRES:
        print("  ✅ Modo: ONLINE (PostgreSQL)")
        print("  URL: Verifique o painel do Render.com")
    else:
        print("  ✅ Modo: LOCAL (SQLite)")
        print()
        print("  Acesse pelo navegador:")
        print("  - Local:    http://localhost:5000")
        print("  - Rede:     http://[SEU-IP]:5000")
        print()
        print("  Login DP:   dp@ghf.com / 123456")

    print()
    print("  Pressione Ctrl+C para parar")
    print("=" * 50)

    app.run(debug=True, host='0.0.0.0', port=5000)
