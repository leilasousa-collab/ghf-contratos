"""
GHF - Modulo de Banco de Dados
Importacao de planilha Excel para SQLite
"""

import os
import sqlite3
import hashlib
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, 'dados', 'contratos.db')

def criar_banco():
    """Cria o banco de dados e tabelas"""
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)

    conn = sqlite3.connect(DATABASE)
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            perfil TEXT NOT NULL CHECK(perfil IN ('dp', 'subgerente', 'gerente_area', 'supervisor')),
            loja TEXT,
            lojas TEXT,
            regional TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS colaboradores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            empresa TEXT,
            nome_empresa TEXT,
            matricula TEXT,
            nome TEXT NOT NULL,
            ddd TEXT,
            celular TEXT,
            funcao TEXT,
            email TEXT,
            data_admissao TEXT,
            vencimento_1 TEXT,
            vencimento_2 TEXT,
            departamento TEXT,
            setor TEXT,
            status_1 TEXT DEFAULT 'pendente' CHECK(status_1 IN ('pendente', 'renovado', 'desligado')),
            status_2 TEXT DEFAULT 'pendente' CHECK(status_2 IN ('pendente', 'efetivado', 'desligado')),
            validacao_gestor_1 TEXT,
            validacao_data_1 TEXT,
            validacao_gestor_2 TEXT,
            validacao_data_2 TEXT,
            alertas_enviados INTEGER DEFAULT 0,
            atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS alertas_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            colaborador_id INTEGER,
            tipo TEXT CHECK(tipo IN ('email', 'whatsapp')),
            periodo INTEGER CHECK(periodo IN (1, 2)),
            data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (colaborador_id) REFERENCES colaboradores(id)
        );

        CREATE TABLE IF NOT EXISTS gestores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cargo TEXT NOT NULL CHECK(cargo IN ('subgerente', 'gerente_area', 'supervisor')),
            loja TEXT,
            lojas TEXT,
            regional TEXT,
            ddd TEXT,
            celular TEXT,
            email TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS config_sistema (
            chave TEXT PRIMARY KEY,
            valor TEXT
        );
    ''')
    conn.commit()
    return conn

def converter_data(data_val):
    """Converte varios formatos de data para YYYY-MM-DD"""
    if data_val is None or str(data_val).strip() == '' or str(data_val).strip() == 'None':
        return None

    # Formato: datetime do openpyxl
    if isinstance(data_val, datetime):
        return data_val.strftime('%Y-%m-%d')

    # Formato: numero do Excel (serial date)
    if isinstance(data_val, (int, float)):
        try:
            # Excel serial date: dias desde 1/1/1900 (com ajuste do bug do Excel)
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)  # Base do Excel
            result_date = base_date + timedelta(days=int(data_val))
            return result_date.strftime('%Y-%m-%d')
        except:
            pass

    # Formato: string
    data_str = str(data_val).strip()

    # Tentar se e um numero (serial date como string)
    try:
        num = float(data_str)
        if num > 30000 and num < 50000:  # Provavel serial date do Excel
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)
            result_date = base_date + timedelta(days=int(num))
            return result_date.strftime('%Y-%m-%d')
    except:
        pass

    # Tentar formatos comuns
    formatos = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']
    for fmt in formatos:
        try:
            return datetime.strptime(data_str, fmt).strftime('%Y-%m-%d')
        except:
            continue
    return None

def importar_excel(caminho_excel, db_conn=None):
    """Importa dados da planilha Excel para o banco de dados.
    Se db_conn for fornecido, usa a conexao existente (PostgreSQL wrapper).
    Caso contrario, cria conexao SQLite local."""
    if not os.path.exists(caminho_excel):
        print(f"ERRO: Arquivo nao encontrado: {caminho_excel}")
        return False

    print(f"Lendo planilha: {caminho_excel}")

    # Detectar formato do arquivo
    extensao = os.path.splitext(caminho_excel)[1].lower()

    if extensao == '.xls':
        # Formato antigo - usar xlrd
        try:
            import xlrd
        except ImportError:
            print("ERRO: Instale xlrd - pip install xlrd")
            return False

        wb = xlrd.open_workbook(caminho_excel)
        ws = wb.sheet_by_index(0)

        # Mapear cabecalhos
        cabecalhos = []
        for cell in ws.row(0):
            cabecalhos.append(str(cell.value).strip().lower() if cell.value else '')

        print(f"Colunas encontradas: {cabecalhos}")

        # Converter rows para formato similar ao openpyxl
        def iter_rows(min_row=1):
            for row_idx in range(min_row, ws.nrows):
                yield [cell.value for cell in ws.row(row_idx)]

    else:
        # Formato .xlsx - usar openpyxl
        try:
            import openpyxl
        except ImportError:
            print("ERRO: Instale openpyxl - pip install openpyxl")
            return False

        wb = openpyxl.load_workbook(caminho_excel, data_only=True)
        ws = wb.active

        # Mapear cabecalhos
        cabecalhos = []
        for cell in ws[1]:
            cabecalhos.append(str(cell.value).strip().lower() if cell.value else '')

        print(f"Colunas encontradas: {cabecalhos}")

        def iter_rows(min_row=2):
            return ws.iter_rows(min_row=min_row, values_only=True)

    print(f"Colunas encontradas: {cabecalhos}")

    # Mapear indices das colunas esperadas
    mapeamento = {
        'empresa': 0,
        'nome da empresa': 1,
        'funcionário': 2,
        'funcionario': 2,
        'nome do funcionário': 3,
        'nome do funcionario': 3,
        'ddd': 4,
        'celular': 5,
        'função': 6,
        'funcao': 6,
        'email': 7,
        'admissão': 8,
        'admissao': 8,
        'vencimento experiência': 9,
        'vencimento experiencia': 9,
        'prorrogação experiência': 10,
        'prorrogacao experiencia': 10,
        'departamento': 11,
        'complemento/setor': 12,
    }

    # Ajustar indices baseado nos cabecalhos reais
    indices = {}
    for nome_col, idx_esperado in mapeamento.items():
        for i, cab in enumerate(cabecalhos):
            if nome_col in cab:
                indices[nome_col] = i
                break

    print(f"Mapeamento: {indices}")

    if db_conn:
        db = db_conn
        close_conn = False
    else:
        conn = criar_banco()
        db = type('DB', (), {
            'execute': lambda self, q, p=None: conn.cursor().execute(q.replace('?', '%s') if p else q, p or ()),
            'fetchall': lambda self, q, p=None: conn.cursor().execute(q.replace('?', '%s') if p else q, p or ()).fetchall(),
            'commit': lambda self: conn.commit(),
            'close': lambda self: conn.close()
        })()
        close_conn = True

    # Manter registros ja validados
    registros_validados = {}
    for row in db.fetchall('SELECT id, matricula, status_1, status_2, validacao_gestor_1, validacao_data_1, validacao_gestor_2, validacao_data_2 FROM colaboradores WHERE status_1 != ? OR status_2 != ?', ('pendente', 'pendente')):
        registros_validados[row[1]] = {
            'status_1': row[2], 'status_2': row[3],
            'validacao_gestor_1': row[4], 'validacao_data_1': row[5],
            'validacao_gestor_2': row[6], 'validacao_data_2': row[7]
        }

    # Limpar dados antigos
    db.execute('DELETE FROM colaboradores')

    importados = 0
    erros = 0

    for row_idx, row in enumerate(iter_rows(min_row=2), start=2):
        try:
            # Pular linhas vazias
            if row[0] is None and row[2] is None:
                continue

            empresa = str(row[indices.get('empresa', 0)] or '') if indices.get('empresa') is not None else ''
            nome_empresa = str(row[indices.get('nome da empresa', 1)] or '') if indices.get('nome da empresa') is not None else ''
            matricula = str(row[indices.get('funcionário', indices.get('funcionario', 2))] or '') if indices.get('funcionário', indices.get('funcionario')) is not None else ''
            nome = str(row[indices.get('nome do funcionário', indices.get('nome do funcionario', 3))] or '') if indices.get('nome do funcionário', indices.get('nome do funcionario')) is not None else ''
            ddd = str(row[indices.get('ddd', 4)] or '') if indices.get('ddd') is not None else ''
            celular = str(row[indices.get('celular', 5)] or '') if indices.get('celular') is not None else ''
            funcao = str(row[indices.get('função', indices.get('funcao', 6))] or '') if indices.get('função', indices.get('funcao')) is not None else ''
            email = str(row[indices.get('email', 7)] or '') if indices.get('email') is not None else ''

            idx_admissao = indices.get('admissão', indices.get('admissao', 8))
            idx_venc1 = indices.get('vencimento experiência', indices.get('vencimento experiencia', 9))
            idx_venc2 = indices.get('prorrogação experiência', indices.get('prorrogacao experiencia', 10))

            data_admissao = converter_data(row[idx_admissao]) if idx_admissao is not None else None
            vencimento_1 = converter_data(row[idx_venc1]) if idx_venc1 is not None else None
            vencimento_2 = converter_data(row[idx_venc2]) if idx_venc2 is not None else None

            idx_dept = indices.get('departamento', 11)
            idx_setor = indices.get('complemento/setor', 12)
            departamento = str(row[idx_dept] or '') if idx_dept is not None else ''
            setor = str(row[idx_setor] or '') if idx_setor is not None else ''

            if not nome:
                continue

            # Recuperar validacoes anteriores
            valid = registros_validados.get(matricula, {})

            db.execute('''
                INSERT INTO colaboradores
                (empresa, nome_empresa, matricula, nome, ddd, celular, funcao, email,
                 data_admissao, vencimento_1, vencimento_2, departamento, setor,
                 status_1, status_2, validacao_gestor_1, validacao_data_1,
                 validacao_gestor_2, validacao_data_2)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                empresa, nome_empresa, matricula, nome, ddd, celular, funcao, email,
                data_admissao, vencimento_1, vencimento_2, departamento, setor,
                valid.get('status_1', 'pendente'),
                valid.get('status_2', 'pendente'),
                valid.get('validacao_gestor_1'),
                valid.get('validacao_data_1'),
                valid.get('validacao_gestor_2'),
                valid.get('validacao_data_2')
            ))
            importados += 1

        except Exception as e:
            erros += 1
            print(f"Erro linha {row_idx}: {e}")

    db.commit()
    if close_conn:
        db.close()
    try:
        wb.close()
    except:
        pass

    print(f"\nResultado: {importados} importados, {erros} erros")
    return True

def criar_usuario_dp():
    """Cria usuario DP padrao se nao existir"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    existe = cursor.execute('SELECT COUNT(*) FROM usuarios WHERE perfil = "dp"').fetchone()[0]
    if existe == 0:
        senha_hash = hashlib.sha256('123456'.encode('utf-8')).hexdigest()
        cursor.execute('''
            INSERT INTO usuarios (nome, email, senha_hash, perfil)
            VALUES (?, ?, ?, ?)
        ''', ('DP Master', 'dp@ghf.com', senha_hash, 'dp'))
        conn.commit()
        print("Usuario DP criado: dp@ghf.com / 123456")
    else:
        print("Usuarios DP ja existem.")

    conn.close()

def listar_colaboradores():
    """Lista todos os colaboradores (debug)"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    colaboradores = conn.execute('SELECT * FROM colaboradores ORDER BY setor, departamento, nome').fetchall()
    conn.close()

    print(f"\nTotal: {len(colaboradores)} colaboradores\n")
    for c in colaboradores[:10]:
        print(f"  {c['nome']} | {c['departamento']} | {c['setor']} | V1: {c['vencimento_1']} | V2: {c['vencimento_2']} | S1: {c['status_1']} | S2: {c['status_2']}")

    if len(colaboradores) > 10:
        print(f"  ... e mais {len(colaboradores) - 10} colaboradores")

def status_sistema():
    """Mostra status atual do sistema"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row

    total = conn.execute('SELECT COUNT(*) as c FROM colaboradores').fetchone()['c']
    p1_pend = conn.execute('SELECT COUNT(*) as c FROM colaboradores WHERE status_1 = "pendente"').fetchone()['c']
    p2_pend = conn.execute('SELECT COUNT(*) as c FROM colaboradores WHERE status_2 = "pendente" AND status_1 != "pendente"').fetchone()['c']
    renovados = conn.execute('SELECT COUNT(*) as c FROM colaboradores WHERE status_1 = "renovado"').fetchone()['c']
    efetivados = conn.execute('SELECT COUNT(*) as c FROM colaboradores WHERE status_2 = "efetivado"').fetchone()['c']
    desligados = conn.execute('SELECT COUNT(*) as c FROM colaboradores WHERE status_1 = "desligado" OR status_2 = "desligado"').fetchone()['c']
    usuarios = conn.execute('SELECT COUNT(*) as c FROM usuarios WHERE ativo = 1').fetchone()['c']

    conn.close()

    print("\n" + "=" * 40)
    print("  STATUS DO SISTEMA GHF")
    print("=" * 40)
    print(f"  Total colaboradores:  {total}")
    print(f"  1o Periodo pendente:  {p1_pend}")
    print(f"  2o Periodo pendente:  {p2_pend}")
    print(f"  Renovados:            {renovados}")
    print(f"  Efetivados:           {efetivados}")
    print(f"  Desligados:           {desligados}")
    print(f"  Usuarios ativos:      {usuarios}")
    print("=" * 40)


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        cmd = sys.argv[1]

        if cmd == 'importar':
            caminho = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE_DIR, 'dados', 'CONTRATO DE EXPERIENCIA.xlsx')
            importar_excel(caminho)
            criar_usuario_dp()

        elif cmd == 'status':
            status_sistema()

        elif cmd == 'listar':
            listar_colaboradores()

        elif cmd == 'usuarios':
            criar_usuario_dp()

        elif cmd == 'reset':
            if os.path.exists(DATABASE):
                os.remove(DATABASE)
                print("Banco removido.")
            criar_banco()
            print("Banco criado.")
        else:
            print("Comandos: importar, status, listar, usuarios, reset")
    else:
        print("Uso: python database.py <comando>")
        print("  importar [arquivo]  - Importa planilha Excel")
        print("  status              - Mostra status do sistema")
        print("  listar              - Lista colaboradores")
        print("  usuarios            - Cria usuario DP padrao")
        print("  reset               - Reseta o banco de dados")
